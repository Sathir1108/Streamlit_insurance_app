import tempfile
import streamlit as st
import os
import json
import pandas as pd
import google.generativeai as genai
from io import BytesIO
import openpyxl
from dotenv import load_dotenv
import re
from datetime import datetime
from streamlit_pdf_viewer import pdf_viewer



load_dotenv()
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')

# Configure Gemini AI
genai.configure(api_key=GEMINI_API_KEY)

# Date and Value Formatting Functions
def standardize_date(date_str):
    """Convert different date formats to DD/MM/YYYY."""
    if not date_str or not isinstance(date_str, str):
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return date_str  # Return original if parsing fails

def format_numeric_value(value):
    """Format numeric values with commas as thousand separators."""
    if not value:
        return ""
    value = re.sub(r'[^0-9.,]', '', str(value))  # Remove non-numeric characters
    try:
        numeric_value = float(value.replace(',', ''))
        return f"{numeric_value:,.0f}"  # Format with commas
    except ValueError:
        return value  # Return original if conversion fails

def save_to_excel(data):
    """Save extracted data to Excel format with formatted dates and values"""
    wb = openpyxl.Workbook()

    # Sheet 1: Policy & Vehicle Details
    ws1 = wb.active
    ws1.title = "Policy & Vehicle Details"
    policy_fields = [
        "Policy_Number", "Full_Name", "NIC_or_Reg_No", "Postal_Address", "Mobile", 
        "Landline", "Email", "preferred_language", "Financial_Interest", 
        "Accident_free_or_other_damages", "Claims_in_Last_3_Years", "Registered_Owner", 
        "Business_Occupation"
    ]
    for col_idx, field in enumerate(policy_fields, start=1):
        ws1.cell(row=1, column=col_idx, value=field)
    for col_idx, field in enumerate(policy_fields, start=1):
        ws1.cell(row=2, column=col_idx, value=str(data.get(field, "")))

    # Sheet 2: Vehicle Information
    ws2 = wb.create_sheet("Vehicle Information")
    vehicle_fields = [
        "Make_Model", "Registration_No", "Chassis_No", "Year_of_Make", 
        "First_Registration_Date", "Country_of_Make", "Fuel_Type", "Cubic_Capacity", 
        "Seating_Capacity", "Vehicle_Registered_As", "Usage_of_Vehicle", 
        "Market_Value", "Extra_Fittings_Value", "Total_Value_Insured"
    ]
    for col_idx, field in enumerate(vehicle_fields, start=1):
        ws2.cell(row=1, column=col_idx, value=field)
    for col_idx, field in enumerate(vehicle_fields, start=1):
        value = data.get(field, "")
        if field in ["First_Registration_Date", "Year_of_Make"]:
            value = standardize_date(value)
        elif field in ["Market_Value", "Extra_Fittings_Value", "Total_Value_Insured"]:
            value = format_numeric_value(value)
        ws2.cell(row=2, column=col_idx, value=str(value))

    # Sheet 3: Insurance Coverage
    ws3 = wb.create_sheet("Insurance Coverage")
    coverage_headers = ["Cover Type", "Amount", "Additional Info"]
    for col_idx, header in enumerate(coverage_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=header)
    covers = data.get("covers", [])
    if isinstance(covers, pd.DataFrame):
        covers = covers.to_dict('records')
    for row_idx, cover in enumerate(covers, start=2):
        ws3.cell(row=row_idx, column=1, value=str(cover.get("Cover Type", "")))
        ws3.cell(row=row_idx, column=2, value=format_numeric_value(cover.get("Amount", "")))
        ws3.cell(row=row_idx, column=3, value=str(cover.get("Additional Info", "")))

    # Sheet 4: Policy & Proposer Details
    ws4 = wb.create_sheet("Policy & Proposer")
    policy_fields = ["Period_From", "Period_To", "Proposer_Date", "Proposer_Signature"]
    for col_idx, field in enumerate(policy_fields, start=1):
        ws4.cell(row=1, column=col_idx, value=field)
    ws4.cell(row=2, column=1, value=standardize_date(data.get("Period_From", "")))
    ws4.cell(row=2, column=2, value=standardize_date(data.get("Period_To", "")))
    ws4.cell(row=2, column=3, value=standardize_date(data.get("proposer_details", {}).get("date", "")))
    ws4.cell(row=2, column=4, value=str(data.get("proposer_details", {}).get("proposer_signature", "")))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def flatten_json(extracted_data):
    """Flatten the extracted JSON and format dates and values"""
    flat_data = {}
    policy = extracted_data.get("Policy & Vehicle Details", {})
    flat_data["Policy_Number"] = policy.get("Policy_Number", "")
    flat_data["Full_Name"] = policy.get("Full_Name", "")
    flat_data["NIC_or_Reg_No"] = policy.get("NIC_or_Reg_No", "")
    flat_data["Postal_Address"] = policy.get("Postal_Address", "")
    flat_data["Mobile"] = policy.get("Mobile", "")
    flat_data["Landline"] = policy.get("Landline", "")
    flat_data["Email"] = policy.get("Email", "")
    flat_data["preferred_language"] = policy.get("preferred_language", "")
    flat_data["Financial_Interest"] = policy.get("Financial_Interest", "")
    flat_data["Accident_free_or_other_damages"] = policy.get("Accident_free_or_other_damages", "")
    flat_data["Claims_in_Last_3_Years"] = policy.get("Claims_in_Last_3_Years", "")
    flat_data["Registered_Owner"] = policy.get("Registered_Owner", "")
    flat_data["Business_Occupation"] = policy.get("Business_Occupation", "")

    vehicle = extracted_data.get("Vehicle Information", {})
    flat_data["Make_Model"] = vehicle.get("Make_Model", "")
    flat_data["Registration_No"] = vehicle.get("Registration_No", "")
    flat_data["Chassis_No"] = vehicle.get("Chassis_No", "")
    flat_data["Year_of_Make"] = standardize_date(vehicle.get("Year_of_Make", ""))
    flat_data["First_Registration_Date"] = standardize_date(vehicle.get("First_Registration_Date", ""))
    flat_data["Country_of_Make"] = vehicle.get("Country_of_Make", "")
    flat_data["Fuel_Type"] = vehicle.get("Fuel_Type", "")
    flat_data["Cubic_Capacity"] = vehicle.get("Cubic_Capacity", "")
    flat_data["Seating_Capacity"] = vehicle.get("Seating_Capacity", "")
    flat_data["Vehicle_Registered_As"] = vehicle.get("Vehicle_Registered_As", "")
    flat_data["Usage_of_Vehicle"] = vehicle.get("Usage_of_Vehicle", "")
    flat_data["Market_Value"] = format_numeric_value(vehicle.get("Market_Value", ""))
    flat_data["Extra_Fittings_Value"] = format_numeric_value(vehicle.get("Extra_Fittings_Value", ""))
    flat_data["Total_Value_Insured"] = format_numeric_value(vehicle.get("Total_Value_Insured", ""))

    covers = extracted_data.get("Insurance Coverage", [])
    flat_data["covers"] = [
        {
            "Cover Type": cover.get("Cover Type", ""),
            "Amount": format_numeric_value(cover.get("Amount", "")),
            "Additional Info": cover.get("Additional Info", "")
        } for cover in covers
    ]

    policy_period = extracted_data.get("Policy & Proposer", {})
    flat_data["Period_From"] = standardize_date(policy_period.get("Period_From", ""))
    flat_data["Period_To"] = standardize_date(policy_period.get("Period_To", ""))
    proposer_signature = policy_period.get("Proposer_Signature", "")
    if proposer_signature and proposer_signature.strip():
        if any(c.isalpha() for c in proposer_signature):
            flat_data["proposer_details"] = {
                "date": standardize_date(policy_period.get("Proposer_Date", "")),
                "proposer_signature": proposer_signature
            }
        else:
            flat_data["proposer_details"] = {
                "date": standardize_date(policy_period.get("Proposer_Date", "")),
                "proposer_signature": "available"
            }
    else:
        flat_data["proposer_details"] = {
            "date": standardize_date(policy_period.get("Proposer_Date", "")),
            "proposer_signature": ""
        }
    return flat_data

def process_document(pdf_bytes):
    """Process PDF document using Gemini AI"""
    
    try:
        import hashlib
        import time

        pdf_hash = hashlib.md5(pdf_bytes).hexdigest()
        cache_key = f"gemini_cache_{pdf_hash}"

        if cache_key in st.session_state:
            return st.session_state[cache_key]


        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
            temp_file.write(pdf_bytes)
            temp_file_path = temp_file.name

        uploaded_file = genai.upload_file(
            path=temp_file_path,
            mime_type="application/pdf"
        )
        model = genai.GenerativeModel('gemini-1.5-flash')

        prompt = (
            "Extract all insurance form fields from the document. Return structured JSON data with: "
            "1. 'Policy & Vehicle Details' including Policy_Number, Full_Name, NIC_or_Reg_No, Postal_Address, Mobile, Landline, Email, preferred_language, Financial_Interest, Accident_free_or_other_damages, Claims_in_Last_3_Years, Registered_Owner, Business_Occupation; "
            "2. 'Vehicle Information' including Make_Model, Registration_No, Chassis_No, Year_of_Make, First_Registration_Date, Country_of_Make, Fuel_Type, Cubic_Capacity, Seating_Capacity, Vehicle_Registered_As, Usage_of_Vehicle, Market_Value, Extra_Fittings_Value, Total_Value_Insured; "
            "3. 'Insurance Coverage' as a list of objects representing all additional coverage options that are ticked, marked, or selected in the form. Each object should include: 'Cover Type' (the name/description of the coverage), 'Amount' (any specified value or limit, if provided, otherwise empty string), and 'Additional Info' (any extra details related to that coverage). Include all ticked/marked coverages from sections like 'Additional Covers'; "
            "4. 'Policy & Proposer' including Period_From, Period_To, Proposer_Date, Proposer_Signature. "
            "For Proposer_Signature, if it contains a readable name, extract the name; if a signature is present but not readable as a name, return 'available'; if no signature is present, return an empty string. "
            "Format all date fields (Year_of_Make, First_Registration_Date, Period_From, Period_To, Proposer_Date) in 'DD/MM/YYYY' format (e.g., '01/01/2018'). "
            "Format all amount fields (Market_Value, Extra_Fittings_Value, Total_Value_Insured, and 'Amount' in Insurance Coverage) with commas as thousand separators (e.g., '4,500,000'). "
            "Ensure the output is valid JSON. If a field is not present or cannot be determined, use an empty string ('') or an empty list ([]) as appropriate."
        )
        
        for attempt in range(2):
            try:
                response = model.generate_content([prompt, uploaded_file])
                break
            except Exception as e:
                if "429" in str(e):
                    time.sleep(60)
                else:
                    raise e

        os.unlink(temp_file_path)

        if response and hasattr(response, 'text') and response.text:
            response_text = response.text.strip()
            if response_text.startswith("```json") and response_text.endswith("```"):
                json_str = response_text[7:-3].strip()
            else:
                json_str = response_text

            def fix_trailing_commas(json_str):
                import re
                json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
                return json_str

            json_str = fix_trailing_commas(json_str)
            try:
                final_data = flatten_json(extracted_data)
                st.session_state[cache_key] = final_data
                return final_data

            except json.JSONDecodeError as e:
                st.error(f"JSON parsing error: {str(e)} - Raw response: {response_text}")
                return None
        else:
            st.error("No valid response text received from Gemini.")
            return None
    except Exception as e:
        st.error(f"Processing error: {str(e)}")
        return None

def main():
    st.set_page_config(layout="wide", page_title="Insurance Document Processor")
    
    st.markdown(
        """
        <h1 style='text-align: center;'>📄 Insurance Document Processing System</h1>
        """,
        unsafe_allow_html=True
    )

    # Initialize session state
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = None
    if 'edited_data' not in st.session_state:
        st.session_state.edited_data = None
    if 'step' not in st.session_state:
        st.session_state.step = 0  
    if 'show_process_button' not in st.session_state:
        st.session_state.show_process_button = True
    if 'excel_file' not in st.session_state:
        st.session_state.excel_file = None  
    if 'show_export_button' not in st.session_state:
        st.session_state.show_export_button = True  
    if 'current_file_name' not in st.session_state:
        st.session_state.current_file_name = None  
    if 'pdf_bytes' not in st.session_state:
        st.session_state.pdf_bytes = None

    # Column widths
    col1, col2 = st.columns([2, 2], gap="large")

    with col1:
        st.markdown("---")
        st.subheader("📤 Document Upload")
        uploaded_file = st.file_uploader("Upload PDF Insurance Document", type=["pdf"])
        
        if uploaded_file:
            new_file_name = uploaded_file.name
            if st.session_state.current_file_name != new_file_name:
                st.session_state.extracted_data = None
                st.session_state.edited_data = None
                st.session_state.step = 0
                st.session_state.show_process_button = True
                st.session_state.excel_file = None
                st.session_state.show_export_button = True
                st.session_state.current_file_name = new_file_name
                st.session_state.pdf_bytes = None
                st.rerun()

            try:
                pdf_bytes = uploaded_file.getvalue()
                st.session_state.pdf_bytes = pdf_bytes

                # Validate PDF size
                if len(pdf_bytes) == 0:
                    st.error("Uploaded PDF is empty.")
                    return

                # Display PDF using streamlit-pdf-viewer
                pdf_viewer(pdf_bytes, width=700, height=1100)

                uploaded_file.seek(0)
            except Exception as e:
                st.error(f"PDF rendering error: {str(e)}")

    with col2:
        st.markdown("---")
        if uploaded_file and st.session_state.show_process_button:
            if st.button("🚀 Process Document"):
                with st.spinner("Analyzing document..."):
                    try:
                        pdf_bytes = st.session_state.pdf_bytes
                        result = process_document(pdf_bytes)
                        if result:
                            st.session_state.extracted_data = result
                            st.session_state.edited_data = result.copy()
                            st.session_state.step = 1
                            st.session_state.show_process_button = False
                            st.success("✅ Document processed successfully!")
                        else:
                            st.error("❌ Failed to extract data from document")
                    except Exception as e:
                        st.error(f"❌ Processing error: {str(e)}")

        if st.session_state.edited_data:  
            if st.session_state.step in [3, 4]:
                st.markdown(
                    """
                    <style>
                    .custom-margin {
                        margin-top: 500px; 
                    }
                    </style>
                    """,
                    unsafe_allow_html=True
                )
                st.markdown('<div class="custom-margin">', unsafe_allow_html=True)
                st.subheader("🔍 Extracted Data")
            else:
                st.subheader("🔍 Extracted Data")

            if st.session_state.step == 1:
                with st.expander("📋 Policy & Vehicle Details", expanded=True):
                    fields = ['Policy_Number', 'Full_Name', 'NIC_or_Reg_No', 'Postal_Address',
                             'Mobile', 'Landline', 'Email', 'preferred_language',
                             'Financial_Interest', 'Accident_free_or_other_damages',
                             'Claims_in_Last_3_Years', 'Registered_Owner', 'Business_Occupation']
                    for field in fields:
                        new_value = st.text_input(
                            field.replace('_', ' ').title(),
                            value=st.session_state.edited_data.get(field, ''),
                            key=f"policy_{field}"
                        )
                        st.session_state.edited_data[field] = new_value

            elif st.session_state.step == 2:
                with st.expander("🚗 Vehicle Information", expanded=True):
                    vehicle_fields = ['Make_Model', 'Registration_No', 'Chassis_No',
                                     'Year_of_Make', 'First_Registration_Date',
                                     'Country_of_Make', 'Fuel_Type', 'Cubic_Capacity',
                                     'Seating_Capacity', 'Vehicle_Registered_As',
                                     'Usage_of_Vehicle', 'Market_Value', 'Extra_Fittings_Value',
                                     'Total_Value_Insured']
                    for field in vehicle_fields:
                        value = st.session_state.edited_data.get(field, '')
                        if field in ['Year_of_Make', 'First_Registration_Date']:
                            value = standardize_date(value)
                        elif field in ['Market_Value', 'Extra_Fittings_Value', 'Total_Value_Insured']:
                            value = format_numeric_value(value)
                        new_value = st.text_input(
                            field.replace('_', ' ').title(),
                            value=value,
                            key=f"vehicle_{field}"
                        )
                        st.session_state.edited_data[field] = new_value

            elif st.session_state.step == 3:
                with st.expander("🛡️ Insurance Coverage", expanded=True):
                    if 'covers' in st.session_state.edited_data:
                        if isinstance(st.session_state.edited_data['covers'], list):
                            coverage_df = pd.DataFrame(st.session_state.edited_data['covers'])
                        else:
                            coverage_df = pd.DataFrame(columns=["Cover Type", "Amount", "Additional Info"])
                        # Format amounts in the coverage dataframe
                        if 'Amount' in coverage_df.columns:
                            coverage_df['Amount'] = coverage_df['Amount'].apply(format_numeric_value)
                        edited_coverage = st.data_editor(
                            coverage_df,
                            column_config={
                                "Cover Type": "Cover Type",
                                "Amount": st.column_config.TextColumn("Amount", help="Enter amount (e.g., 4500000)"),
                                "Additional Info": "Additional Info"
                            },
                            num_rows="dynamic",
                            use_container_width=True,
                            key="coverage_editor"
                        )
                        st.session_state.edited_data['covers'] = edited_coverage.to_dict('records')

            elif st.session_state.step == 4:
                with st.expander("📅 Policy & Proposer Details", expanded=True):
                    new_period_from = st.text_input(
                        "Period From",
                        value=standardize_date(st.session_state.edited_data.get('Period_From', '')),
                        key="period_from"
                    )
                    new_period_to = st.text_input(
                        "Period To",
                        value=standardize_date(st.session_state.edited_data.get('Period_To', '')),
                        key="period_to"
                    )
                    new_proposer_date = st.text_input(
                        "Proposer Date",
                        value=standardize_date(st.session_state.edited_data.get('proposer_details', {}).get('date', '')),
                        key="proposer_date"
                    )
                    new_proposer_signature = st.text_input(
                        "Proposer Signature",
                        value=st.session_state.edited_data.get('proposer_details', {}).get('proposer_signature', ''),
                        key="proposer_signature"
                    )
                    st.session_state.edited_data['Period_From'] = new_period_from
                    st.session_state.edited_data['Period_To'] = new_period_to
                    st.session_state.edited_data['proposer_details'] = {
                        "date": new_proposer_date,
                        "proposer_signature": new_proposer_signature if any(c.isalpha() for c in new_proposer_signature) else "available" if new_proposer_signature else ""
                    }

            if st.session_state.step in [3, 4]:
                st.markdown('</div>', unsafe_allow_html=True)

            col_left, col_space, col_right = st.columns([3, 8, 2])
            with col_left:
                if st.session_state.step > 1:
                    if st.button("⬅️ Previous", key="prev_btn"):
                        st.session_state.step -= 1
                        st.rerun()
            with col_right:
                if st.session_state.step < 4:
                    if st.button("➡️ Next", key="next_btn"):
                        st.session_state.step += 1
                        st.rerun()

            if st.session_state.step == 4:
                st.markdown("---")
                button_placeholder = st.empty()

                if st.session_state.show_export_button:
                    if button_placeholder.button("💾 Export to Excel", key="export_btn"):
                        with st.spinner("Generating Excel file..."):
                            try:
                                export_data = st.session_state.edited_data.copy()
                                if 'covers' not in export_data or export_data['covers'] is None:
                                    export_data['covers'] = []
                                elif isinstance(export_data['covers'], pd.DataFrame):
                                    export_data['covers'] = export_data['covers'].to_dict('records')
                                excel_file = save_to_excel(export_data)
                                st.session_state.excel_file = excel_file.getvalue()
                                st.session_state.show_export_button = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Export error: {str(e)}")
                else:
                    button_placeholder.download_button(
                        label="⬇️ Download Excel File",
                        data=st.session_state.excel_file,
                        file_name="insurance_details.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_btn"
                    )

if __name__ == "__main__":
    main()

